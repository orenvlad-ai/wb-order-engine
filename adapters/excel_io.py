import io
from datetime import date, datetime
from typing import Tuple, List, Dict, Any, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pydantic import ValidationError

from engine.models import SkuInput, InTransitItem, Recommendation
from engine.calc import calculate

REQUIRED_INPUT_COLS = ["sku", "stock_ff", "stock_mp", "plan_sales_per_day"]
OPTIONAL_INPUT_COLS = ["safety_stock_ff", "safety_stock_mp"]
REQUIRED_INTRANSIT_COLS = ["sku", "qty", "eta_cn_msk"]
SETTINGS_SHEET_NAME = "Настройки заказа"

INPUT_SHEET_NAMES = ("Ввод данных", "Ввод", "Input")
INTRANSIT_SHEET_NAMES = ("InTransit", "Товары в пути")

# Отображения колонок листа Recommendations
RECOMMENDATION_COLUMN_ALIASES = {
    "sku":             "Артикул",
    "order_qty":       "Рек.заказ",
    "current_plan":    "Тек.план\nшт/день",
    "reco_before_1p":  "Рек\nдо 1П",
    "stock_before_1":  "Ост. до\n1П",
    "stock_after_1":   "Ост. после\n1П",
    "reco_before_2p":  "Рек\nдо 2П",
    "stock_before_2":  "Ост. до\n2П",
    "stock_after_2":   "Ост. после\n2П",
    "reco_before_3p":  "Рек\nдо 3П",
    "stock_before_3":  "Ост. до\n3П",
    "stock_after_3":   "Ост. после\n3П",
    "reco_before_po":  "Рек\nдо РП",
    "eoh":             "Ост. до\nРП",
    "stock_after_po":  "Ост. после\nРП",
    "eop_first":       "Ост. после\n1П",  # унификация
    "H_days":          "Горизонт\nпрогноза",
    "coverage":        "Покрытие",
    "inbound":         "В пути",
    "onhand":          "Ост. на\nруках",
    "demand_H":        "Спрос за\nгоризонт",
    "target":          "Цель",
    "shortage":        "Нехватка",
    "moq_step":        "Кратность\n(MOQ)",
    "stock_status":    "Статус",
    "algo_version":    "Версия\nалгоритма",
    "oos_threshold":   "_thr",   # служебный столбец, будет скрыт
}

RECOMMENDATION_DISPLAY_TO_INTERNAL: Dict[str, str] = {}
for internal, display in RECOMMENDATION_COLUMN_ALIASES.items():
    RECOMMENDATION_DISPLAY_TO_INTERNAL.setdefault(display, internal)

# Сопоставление русских заголовков с внутренними ключами
INPUT_COLUMN_ALIASES = {
    "Артикул": "sku",
    "Остаток ФФ": "stock_ff",
    "Остаток МП": "stock_mp",
    "План, шт/день": "plan_sales_per_day",
    "Несниж. остаток ФФ": "safety_stock_ff",
    "Несниж. остаток МП": "safety_stock_mp",
}

INTRANSIT_COLUMN_ALIASES = {
    "Артикул": "sku",
    "Кол-во": "qty",
    "План. приб. на ФФ": "eta_cn_msk",
}

SETTINGS_COLUMN_ALIASES = {
    "Произв., дней": "prod_lead_time_days",
    "Китай→МСК, дней": "lead_time_cn_msk",
    "МСК→МП, дней": "lead_time_msk_mp",
    "Кратность (MOQ)": "moq_step_default",
    "Порог несниж. МП при OOS, %": "oos_safety_mp_pct",
    "Дефолт. несниж. ФФ": "safety_stock_ff_default",
    "Дефолт. несниж. МП": "safety_stock_mp_default",
    "Коэф. несн. ФФ": "safety_stock_ff_coeff",
    "Коэф. несн. МП": "safety_stock_mp_coeff",
}

# Отображение внутренних имён обратно в русские заголовки для сообщений об ошибках
INPUT_COLUMN_DISPLAY = {
    "sku": "Артикул",
    "stock_ff": "Остаток ФФ",
    "stock_mp": "Остаток МП",
    "plan_sales_per_day": "План, шт/день",
    "safety_stock_ff": "Несниж. остаток ФФ",
    "safety_stock_mp": "Несниж. остаток МП",
}

INTRANSIT_COLUMN_DISPLAY = {
    "sku": "Артикул",
    "qty": "Кол-во",
    "eta_cn_msk": "План. приб. на ФФ",
}

SETTINGS_COLUMN_DISPLAY = {
    "prod_lead_time_days": "Произв., дней",
    "lead_time_cn_msk": "Китай→МСК, дней",
    "lead_time_msk_mp": "МСК→МП, дней",
    "moq_step_default": "Кратность (MOQ)",
    "oos_safety_mp_pct": "Порог несниж. МП при OOS, %",
    "safety_stock_ff_default": "Дефолт. несниж. ФФ",
    "safety_stock_mp_default": "Дефолт. несниж. МП",
    "safety_stock_ff_coeff": "Коэф. несн. ФФ",
    "safety_stock_mp_coeff": "Коэф. несн. МП",
}
REQUIRED_SETTINGS_COLS = [
    "prod_lead_time_days",
    "lead_time_cn_msk",
    "lead_time_msk_mp",
    "moq_step_default",
    "oos_safety_mp_pct",
]
OPTIONAL_SETTINGS_COLS = [
    "safety_stock_ff_default",
    "safety_stock_mp_default",
]
OPTIONAL_SETTINGS_FLOAT_COLS = [
    "safety_stock_ff_coeff",
    "safety_stock_mp_coeff",
]

class BadTemplateError(Exception): ...


def _ensure_columns(
    df: pd.DataFrame,
    required: List[str],
    sheet: str,
    display_names: Optional[Dict[str, str]] = None,
):
    missing = [c for c in required if c not in df.columns]
    if missing:
        readable = [display_names.get(c, c) for c in missing] if display_names else missing
        raise BadTemplateError(
            f"На листе '{sheet}' отсутствуют обязательные колонки: {', '.join(readable)}."
        )


def _is_blank(value: Any) -> bool:
    if pd.isna(value):
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _parse_int(value: Any, *, sheet: str, column: str, sku: Optional[str] = None) -> int:
    if _is_blank(value):
        target = f" для SKU '{sku}'" if sku else ""
        raise BadTemplateError(
            f"На листе '{sheet}' колонка '{column}'{target} не заполнена."
        )
    try:
        return int(value)
    except (TypeError, ValueError):
        target = f" для SKU '{sku}'" if sku else ""
        raise BadTemplateError(
            f"На листе '{sheet}' колонка '{column}'{target} должна содержать целое число."
        )


def _parse_float(value: Any, *, sheet: str, column: str, sku: str) -> float:
    if _is_blank(value):
        raise BadTemplateError(
            f"На листе '{sheet}' колонка '{column}' для SKU '{sku}' не заполнена."
        )
    try:
        return float(value)
    except (TypeError, ValueError):
        raise BadTemplateError(
            f"На листе '{sheet}' колонка '{column}' для SKU '{sku}' должна содержать число."
        )


def _read_settings(df_settings: pd.DataFrame) -> Dict[str, Any]:
    # Считываем первую заполненную строку с общими параметрами заказа
    _ensure_columns(
        df_settings,
        REQUIRED_SETTINGS_COLS,
        SETTINGS_SHEET_NAME,
        SETTINGS_COLUMN_DISPLAY,
    )
    df_clean = df_settings.dropna(how="all")
    if df_clean.empty:
        raise BadTemplateError(
            "Лист 'Настройки заказа' пуст. Заполни строку с параметрами по умолчанию."
        )

    row = df_clean.iloc[0].to_dict()
    settings: Dict[str, Any] = {}
    for key in REQUIRED_SETTINGS_COLS:
        value = row.get(key)
        if key == "oos_safety_mp_pct":
            value = 5 if _is_blank(value) else value
            settings[key] = _parse_float(
                value,
                sheet=SETTINGS_SHEET_NAME,
                column=SETTINGS_COLUMN_DISPLAY.get(key, key),
                sku="*",
            )
        else:
            settings[key] = _parse_int(
                value,
                sheet=SETTINGS_SHEET_NAME,
                column=SETTINGS_COLUMN_DISPLAY.get(key, key),
            )
    for key in OPTIONAL_SETTINGS_COLS:
        if key in df_settings.columns:
            value = row.get(key)
            settings[key] = _parse_int(
                value,
                sheet=SETTINGS_SHEET_NAME,
                column=SETTINGS_COLUMN_DISPLAY.get(key, key),
            )
        else:
            settings[key] = 0
    for key in OPTIONAL_SETTINGS_FLOAT_COLS:
        if key in df_settings.columns:
            value = row.get(key)
            settings[key] = _parse_float(
                value,
                sheet=SETTINGS_SHEET_NAME,
                column=SETTINGS_COLUMN_DISPLAY.get(key, key),
                sku="*",
            )
        else:
            settings[key] = 1.0
    return settings

def read_input(xlsx_bytes: bytes) -> Tuple[List[SkuInput], List[InTransitItem]]:
    xl = pd.ExcelFile(io.BytesIO(xlsx_bytes))
    input_sheet_name = next((name for name in INPUT_SHEET_NAMES if name in xl.sheet_names), None)
    if input_sheet_name is None:
        raise BadTemplateError("В файле нет листа 'Input', 'Ввод данных' или 'Ввод'.")

    df_in = pd.read_excel(xl, input_sheet_name)
    df_in.rename(columns=INPUT_COLUMN_ALIASES, inplace=True)
    df_in = df_in.where(pd.notna(df_in), None)

    # --- Нормализация SKU: удаляем неразрывные пробелы, выравниваем регистр, заменяем длинные дефисы ---
    def _normalize_sku(s: Any) -> str:
        if not isinstance(s, str):
            return s
        # 1) базовая чистка: обрезка, неразрывные пробелы, длинные дефисы, регистр
        s = (
            s.strip()
            .replace("\xa0", " ")
            .replace("–", "-")
            .replace("—", "-")
            .lower()
        )
        # 2) убрать пробелы вокруг слэша: " / ", " /", "/ " -> "/"
        s = s.replace(" / ", "/").replace(" /", "/").replace("/ ", "/")
        # 3) схлопнуть лишние пробелы (в т.ч. табы/множественные)
        s = " ".join(s.split())
        return s

    df_in["sku"] = df_in["sku"].apply(_normalize_sku)

    try:
        df_settings = pd.read_excel(xl, SETTINGS_SHEET_NAME)
    except ValueError as e:
        raise BadTemplateError("В файле нет листа 'Настройки заказа'.") from e

    df_settings.rename(columns=SETTINGS_COLUMN_ALIASES, inplace=True)
    df_settings = df_settings.where(pd.notna(df_settings), None)

    # --- Читаем лист "Товары в пути" и нормализуем SKU + даты ETA ---
    transit_sheet_name = next(
        (name for name in INTRANSIT_SHEET_NAMES if name in xl.sheet_names),
        None,
    )
    if transit_sheet_name:
        df_tr = pd.read_excel(xl, transit_sheet_name)
        df_tr.rename(columns=INTRANSIT_COLUMN_ALIASES, inplace=True)
        df_tr = df_tr.where(pd.notna(df_tr), None)
        _ensure_columns(
            df_tr,
            REQUIRED_INTRANSIT_COLS,
            transit_sheet_name,
            INTRANSIT_COLUMN_DISPLAY,
        )

        # Нормализация артикулов
        if "sku" in df_tr.columns:
            df_tr["sku"] = df_tr["sku"].apply(_normalize_sku)

        # Приведение ETA к формату datetime.date
        if "eta_cn_msk" in df_tr.columns:
            df_tr["eta_cn_msk"] = pd.to_datetime(df_tr["eta_cn_msk"], errors="coerce").dt.date

        # Удаляем пустые строки без артикулов
        df_tr = df_tr[df_tr["sku"].notna()]
    else:
        df_tr = pd.DataFrame(columns=["sku", "qty", "eta_cn_msk"])

    _ensure_columns(
        df_in,
        REQUIRED_INPUT_COLS,
        input_sheet_name,
        INPUT_COLUMN_DISPLAY,
    )
    for col in OPTIONAL_INPUT_COLS:
        if col not in df_in.columns:
            df_in[col] = None

    settings = _read_settings(df_settings)

    items: List[SkuInput] = []
    moq_step_default = settings["moq_step_default"]
    safety_stock_ff_default = settings["safety_stock_ff_default"]
    safety_stock_mp_default = settings["safety_stock_mp_default"]
    oos_safety_mp_pct = float(settings.get("oos_safety_mp_pct", 5))
    prod_lead_time_days = settings["prod_lead_time_days"]
    lead_time_cn_msk = settings["lead_time_cn_msk"]
    lead_time_msk_mp = settings["lead_time_msk_mp"]

    for r in df_in.to_dict("records"):
        if all(_is_blank(v) for v in r.values()):
            continue
        sku = str(r.get("sku") or "").strip()
        if not sku:
            raise BadTemplateError(
                f"На листе '{input_sheet_name}' есть строка без SKU. Удали пустые строки."
            )
        try:
            stock_ff = _parse_int(
                r.get("stock_ff"),
                sheet=input_sheet_name,
                column=INPUT_COLUMN_DISPLAY["stock_ff"],
                sku=sku,
            )
            stock_mp = _parse_int(
                r.get("stock_mp"),
                sheet=input_sheet_name,
                column=INPUT_COLUMN_DISPLAY["stock_mp"],
                sku=sku,
            )
            plan_sales_per_day = _parse_float(
                r.get("plan_sales_per_day"),
                sheet=input_sheet_name,
                column=INPUT_COLUMN_DISPLAY["plan_sales_per_day"],
                sku=sku,
            )

            raw_mp = r.get("safety_stock_mp")
            if _is_blank(raw_mp):
                safety_stock_mp = safety_stock_mp_default  # пусто → берём дефолт из настроек
            else:
                safety_stock_mp = _parse_int(
                    raw_mp,
                    sheet=input_sheet_name,
                    column=INPUT_COLUMN_DISPLAY["safety_stock_mp"],
                    sku=sku,
                )

            raw_ff = r.get("safety_stock_ff")
            if _is_blank(raw_ff):
                safety_stock_ff = safety_stock_ff_default  # пусто → берём дефолт из настроек
            else:
                safety_stock_ff = _parse_int(
                    raw_ff,
                    sheet=input_sheet_name,
                    column=INPUT_COLUMN_DISPLAY["safety_stock_ff"],
                    sku=sku,
                )

            items.append(SkuInput(
                sku=sku,
                stock_ff=stock_ff,
                stock_mp=stock_mp,
                plan_sales_per_day=plan_sales_per_day,
                prod_lead_time_days=prod_lead_time_days,
                lead_time_cn_msk=lead_time_cn_msk,
                lead_time_msk_mp=lead_time_msk_mp,
                oos_safety_mp_pct=oos_safety_mp_pct,
                safety_stock_mp=safety_stock_mp,
                safety_stock_ff=safety_stock_ff,
                moq_step=moq_step_default,
            ))
        except (ValueError, ValidationError) as e:
            raise BadTemplateError(
                f"На листе '{input_sheet_name}' строка для SKU '{sku}' содержит неверные данные."
            ) from e

    trans: List[InTransitItem] = []
    for r in df_tr.to_dict("records"):
        if r.get("sku") is None:
            continue
        try:
            sku_norm = _normalize_sku(r.get("sku"))
            eta_val = r.get("eta_cn_msk")
            if not eta_val or pd.isna(eta_val):
                raise ValueError("Пустая или некорректная дата ETA")
            trans.append(
                InTransitItem(
                    sku=sku_norm,
                    qty=int(r.get("qty", 0) or 0),
                    eta_cn_msk=eta_val,
                )
            )
        except (ValueError, ValidationError) as e:
            raise BadTemplateError(
                f"На листе '{transit_sheet_name or INTRANSIT_SHEET_NAMES[0]}' строка для SKU "
                f"'{r.get('sku')}' имеет неверные количество или дату."
            ) from e

    return items, trans

# ---------- Форматирование Recommendations ----------

_HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
_RISK_FILL   = PatternFill(start_color="FFFFE5E5", end_color="FFFFE5E5", fill_type="solid")
_RECO_FILL   = PatternFill(start_color="FFFFFFCC", end_color="FFFFFFCC", fill_type="solid")  # очень светло-желтый
_PLAN_FILL   = PatternFill(start_color="FFE6F7E6", end_color="FFE6F7E6", fill_type="solid")  # очень светло-зеленый
_BOLD = Font(bold=True)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_THIN = Side(border_style="thin", color="FFBFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Порядок колонок (берём те, что реально есть в данных)
_ORDER = [
    "sku", "order_qty", "onhand",
    "current_plan",
    "reco_before_1p", "stock_before_1", "stock_after_1",
    "reco_before_2p", "stock_before_2", "stock_after_2",
    "reco_before_3p", "stock_before_3", "stock_after_3",
    "reco_before_po", "eoh", "stock_after_po",
    "H_days",
    "coverage", "inbound",
    "demand_H", "target", "shortage",
    "moq_step", "stock_status", "algo_version", "oos_threshold",
]


_HEADER_TIPS: Dict[str, str] = {
    "sku": "Артикул — код товара/модель.",
    "order_qty": "Рекомендуемый заказ, шт — что заказать сейчас (округлено до кратности).",
    "stock_status": "Статус запаса — минимум запаса на горизонте при текущем плане: ✅ хватает / ⚠️ не хватает.",
    "current_plan": "Текущий план, шт/день — фактический план из «Ввод данных».",
    "inbound": "В пути, шт — сумма поставок, что успеют на МП до (сегодня+H).",
    "onhand": "Остаток на руках, шт — запасы на момент ввода = Остаток ФФ + Остаток МП.",
    "demand_H": "Спрос за горизонт, шт — продажи за H при текущем плане.",
    "eoh": "Остаток за день до прихода расчётной партии, без order_qty.",
    "stock_before_1": "Ост. до 1-й поставки — остаток накануне первой поставки по текущему плану.",
    "stock_after_1": "Ост. после 1-й поставки — остаток сразу после первой поставки.",
    "stock_before_2": "Ост. до 2-й поставки — остаток накануне второй поставки (если есть).",
    "stock_after_2": "Ост. после 2-й поставки — остаток сразу после второй поставки (если есть).",
    "stock_before_3": "Ост. до 3-й поставки — остаток накануне третьей поставки (если есть).",
    "stock_after_3": "Ост. после 3-й поставки — остаток сразу после третьей поставки (если есть).",
    "stock_after_po": "Ост. после расчётной партии — остаток после прихода расчётной партии (при ненулевом заказе).",
    "H_days": "Горизонт прогноза, дней — H = Произв. + Китай→МСК + МСК→МП.",
    "coverage": "Покрытие, шт — доступный объём за H = Остаток на руках + В пути.",
    "target": "Цель, шт — запас, нужный на конец H = Спрос за горизонт + Неснижаемые (ФФ+МП).",
    "shortage": "Нехватка, шт — max(Цель − Покрытие, 0).",
    "moq_step": "Кратность заказа (MOQ) — шаг округления заказа.",
    "algo_version": "Версия алгоритма — версия логики расчёта.",
}

def _order_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = [c for c in _ORDER if c in df.columns] + [c for c in df.columns if c not in _ORDER]
    return df[cols]

def _find_col_idx_by_internal(ws, internal_key: str) -> int | None:
    header_rows = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not header_rows:
        return None
    for idx, value in enumerate(header_rows[0], start=1):
        if value == internal_key:
            return idx
        if RECOMMENDATION_DISPLAY_TO_INTERNAL.get(value) == internal_key:
            return idx
    return None


def _apply_formats_localized(ws):
    idx_order  = _find_col_idx_by_internal(ws, "order_qty")
    idx_status = _find_col_idx_by_internal(ws, "stock_status")
    idx_short  = _find_col_idx_by_internal(ws, "shortage")
    idx_cov    = _find_col_idx_by_internal(ws, "coverage")
    idx_sku    = _find_col_idx_by_internal(ws, "sku")
    idx_thr    = _find_col_idx_by_internal(ws, "oos_threshold")
    idx_plan   = _find_col_idx_by_internal(ws, "current_plan")
    risk_cols = [
        _find_col_idx_by_internal(ws, "stock_before_1"),
        _find_col_idx_by_internal(ws, "stock_after_1"),
        _find_col_idx_by_internal(ws, "stock_before_2"),
        _find_col_idx_by_internal(ws, "stock_after_2"),
        _find_col_idx_by_internal(ws, "stock_before_3"),
        _find_col_idx_by_internal(ws, "stock_after_3"),
        _find_col_idx_by_internal(ws, "eoh"),
        _find_col_idx_by_internal(ws, "stock_before_po"),
    ]
    risk_cols = [c for c in risk_cols if c]

    reco_cols = [
        _find_col_idx_by_internal(ws, "reco_before_1p"),
        _find_col_idx_by_internal(ws, "reco_before_2p"),
        _find_col_idx_by_internal(ws, "reco_before_3p"),
        _find_col_idx_by_internal(ws, "reco_before_po"),
    ]
    reco_cols = [c for c in reco_cols if c]

    _apply_formats(
        ws,
        idx_order=idx_order,
        idx_status=idx_status,
        idx_short=idx_short,
        idx_cov=idx_cov,
    )

    if idx_thr:
        ws.column_dimensions[get_column_letter(idx_thr)].hidden = True

    # Точечная подсветка всех ячеек ниже порога (красный)
    if idx_status and idx_sku and idx_thr and risk_cols:
        for r in range(2, ws.max_row + 1):
            status_val = str(ws.cell(r, idx_status).value or "")
            if "Не хватает" not in status_val:
                continue
            thr_cell = ws.cell(r, idx_thr).value
            try:
                thr = float(thr_cell) if thr_cell is not None else None
            except Exception:
                thr = None
            if thr is None:
                continue
            breach_found = False
            for c in risk_cols:
                try:
                    val = ws.cell(r, c).value
                    if val is not None and float(val) < thr:
                        ws.cell(r, c).fill = _RISK_FILL
                        breach_found = True
                except Exception:
                    continue
            if breach_found:
                ws.cell(r, idx_sku).fill = _RISK_FILL

    if reco_cols and idx_plan:
        for r in range(2, ws.max_row + 1):
            any_reco = False
            for c in reco_cols:
                val = ws.cell(r, c).value
                if val is None or val == "–":
                    continue
                try:
                    float(val)
                except Exception:
                    continue
                ws.cell(r, c).fill = _RECO_FILL
                any_reco = True
            if any_reco:
                ws.cell(r, idx_plan).fill = _PLAN_FILL


def _apply_formats(
    ws,
    *,
    idx_order: int | None = None,
    idx_status: int | None = None,
    idx_short: int | None = None,
    idx_cov: int | None = None,
):
    header_internal: Dict[int, str] = {}

    # Шапка
    for idx, cell in enumerate(ws[1], start=1):
        internal_name = RECOMMENDATION_DISPLAY_TO_INTERNAL.get(cell.value, cell.value)
        header_internal[idx] = internal_name
        cell.font = _BOLD
        cell.alignment = _CENTER
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        if cell.comment:
            cell.comment = None
        tip_text = _HEADER_TIPS.get(internal_name)
        if tip_text:
            cell.comment = Comment(tip_text, "WB Engine")

    for col_idx, name in (
        (idx_order, "order_qty"),
        (idx_status, "stock_status"),
        (idx_short, "shortage"),
        (idx_cov, "coverage"),
    ):
        if col_idx and header_internal.get(col_idx) != name:
            header_internal[col_idx] = name

    # Форматы чисел и выравнивание данных
    int_like_right = {
        "H_days", "inbound", "onhand", "coverage", "target", "shortage",
        "moq_step", "order_qty", "current_plan", "demand_H", "eop_first", "eoh"
    }
    int_like_center = {
        "reco_before_1p", "reco_before_2p", "reco_before_3p", "reco_before_po"
    }
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            header_value = header_internal.get(cell.column)
            cell.border = _BORDER
            if header_value in int_like_center:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif header_value in int_like_right:
                cell.number_format = "0"
                cell.alignment = Alignment(horizontal="right", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # (удалено) глобальное окрашивание всей строки; подсветку делаем в _apply_formats_localized

    # Автоширина с учётом переносов (wrap_text=True для шапок уже выставлен выше)
    widths = {}
    for r in ws.iter_rows(values_only=True):
        for idx, v in enumerate(r, start=1):
            s = "" if v is None else str(v)
            l = max((len(part) for part in s.replace("\r\n", "\n").split("\n")), default=0)
            widths[idx] = max(widths.get(idx, 0), l)
    for idx, w in widths.items():
        col = get_column_letter(idx)
        # Узкие капы для «Ост. …», чуть шире — для длинных заголовков
        hdr = ws.cell(1, idx).value or ""
        cap = 22
        if "Ост." in hdr or "В пути" in hdr or "Покрытие" in hdr:
            cap = 16
        if "Версия" in hdr:
            cap = 18
        ws.column_dimensions[col].width = min(max(w + 2, 8), cap)


def _auto_width_all(ws):
    """Устанавливает ширину колонок по максимальной длине контента."""

    widths = {}
    for row in ws.iter_rows(values_only=True):
        for idx, value in enumerate(row, start=1):
            if value is None:
                length = 0
            else:
                text = str(value).replace("\r\n", "\n")
                parts = text.split("\n") if text else []
                length = max((len(part) for part in parts), default=len(text))
            widths[idx] = max(widths.get(idx, 0), length)
    for idx, width in widths.items():
        column = get_column_letter(idx)
        ws.column_dimensions[column].width = min(max(width + 2, 8), 40)


def generate_input_template() -> io.BytesIO:
    wb = Workbook()

    input_headers = [
        "Артикул",
        "Остаток ФФ",
        "Остаток МП",
        "План, шт/день",
        "Несниж. остаток ФФ",
        "Несниж. остаток МП",
    ]
    settings_headers = [
        "Произв., дней",
        "Китай→МСК, дней",
        "МСК→МП, дней",
        "Кратность (MOQ)",
        "Порог несниж. МП при OOS, %",
        "Коэф. несн. ФФ",
        "Коэф. несн. МП",
    ]

    ws_input = wb.active
    ws_input.title = "Ввод данных"
    ws_input.append(input_headers)
    for cell in ws_input[1]:
        cell.font = _BOLD

    ws_input.append([
        "SKU123",
        900,
        650,
        14.5,
        "=CEILING(D2*'Настройки заказа'!$F$2, 'Настройки заказа'!$D$2)",
        "=CEILING(D2*'Настройки заказа'!$G$2, 'Настройки заказа'!$D$2)",
    ])
    _auto_width_all(ws_input)

    ws_settings = wb.create_sheet(SETTINGS_SHEET_NAME)
    ws_settings.append(settings_headers)
    for cell in ws_settings[1]:
        cell.font = _BOLD

    ws_settings.append([
        15,
        25,
        10,
        250,
        5,
        10,
        20,
    ])
    _auto_width_all(ws_settings)

    ws_transit = wb.create_sheet("Товары в пути")
    ws_transit.append(["Артикул", "Кол-во", "План. приб. на ФФ"])
    for cell in ws_transit[1]:
        cell.font = _BOLD

    ws_transit.append([
        "SKU123",
        120,
        "2025-11-01",
    ])
    _auto_width_all(ws_transit)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def build_output(xlsx_in: bytes, recs: List[Recommendation]) -> bytes:
    # Конвертируем в DataFrame и отсортируем колонки
    df_rec = pd.DataFrame([r.model_dump() for r in recs])
    if not df_rec.empty:
        df_rec = _order_columns(df_rec)

    in_buf = io.BytesIO(xlsx_in)
    out_buf = io.BytesIO()
    with pd.ExcelWriter(out_buf, engine="openpyxl") as w:
        # 1) Формируем лист "Заказ на фабрику" (первый) — устойчивый к пустым/нестандартным данным
        if not df_rec.empty and "order_qty" in df_rec.columns:
            mask = pd.to_numeric(df_rec["order_qty"], errors="coerce").fillna(0).gt(0)
            columns = [col for col in ("sku", "order_qty") if col in df_rec.columns]
            df_factory = df_rec.loc[mask, columns]
        else:
            df_factory = pd.DataFrame(columns=["sku", "order_qty"])
        df_factory = df_factory.rename(
            columns={"sku": "Артикул\n货号", "order_qty": "Заказ, штук\n数量（件）"}
        )
        df_factory = df_factory.reindex(columns=["Артикул\n货号", "Заказ, штук\n数量（件）"])
        df_factory.to_excel(w, sheet_name="Заказ на фабрику", index=False)
        ws_factory = w.book["Заказ на фабрику"]
        ws_factory.row_dimensions[1].height = 32
        for cell in ws_factory[1]:
            cell.alignment = _CENTER
        _auto_width_all(ws_factory)

        # 2) Пишем лист "Рекомендации": подтягиваем current_plan и onhand из входного листа
        df_out = df_rec.copy()
        if not df_out.empty:
            def _normalize_sku(value: Any) -> Any:
                if not isinstance(value, str):
                    return value
                normalized = (
                    value.strip()
                    .replace("\xa0", " ")
                    .replace("–", "-")
                    .replace("—", "-")
                    .lower()
                )
                normalized = (
                    normalized.replace(" / ", "/")
                    .replace(" /", "/")
                    .replace("/ ", "/")
                )
                return " ".join(normalized.split())

            plan_map = None
            onhand_map = None
            try:
                with pd.ExcelFile(in_buf) as xl_in:
                    input_sheet = next((name for name in INPUT_SHEET_NAMES if name in xl_in.sheet_names), None)
                    if input_sheet:
                        df_in = pd.read_excel(xl_in, input_sheet)
                        df_in = df_in.rename(columns=INPUT_COLUMN_ALIASES)
                        if "sku" in df_in.columns:
                            df_in["sku"] = df_in["sku"].apply(_normalize_sku)
                            for col in ("stock_ff", "stock_mp", "plan_sales_per_day"):
                                if col in df_in.columns:
                                    df_in[col] = pd.to_numeric(df_in[col], errors="coerce")
                            agg_spec = {}
                            if "plan_sales_per_day" in df_in.columns:
                                agg_spec["plan_sales_per_day"] = "max"
                            for col in ("stock_ff", "stock_mp"):
                                if col in df_in.columns:
                                    agg_spec[col] = "sum"
                            if agg_spec:
                                grp = df_in.groupby("sku", as_index=True).agg(agg_spec)
                                if "plan_sales_per_day" in grp.columns:
                                    plan_map = grp["plan_sales_per_day"]
                                if "stock_ff" in grp.columns or "stock_mp" in grp.columns:
                                    stock_ff = grp["stock_ff"] if "stock_ff" in grp.columns else pd.Series(0, index=grp.index)
                                    stock_mp = grp["stock_mp"] if "stock_mp" in grp.columns else pd.Series(0, index=grp.index)
                                    onhand_map = stock_ff.fillna(0) + stock_mp.fillna(0)
            except Exception:
                plan_map = None
                onhand_map = None
            finally:
                in_buf.seek(0)

            if plan_map is not None:
                df_out["current_plan"] = df_out["sku"].map(plan_map)
            elif "current_plan" not in df_out.columns:
                df_out["current_plan"] = pd.Series([None] * len(df_out), index=df_out.index)

            if onhand_map is not None:
                df_out["onhand"] = df_out["sku"].map(onhand_map).fillna(0)
            elif "onhand" not in df_out.columns:
                df_out["onhand"] = 0

            diag_cols = [
                "reco_before_1p", "stock_before_1", "stock_after_1",
                "reco_before_2p", "stock_before_2", "stock_after_2",
                "reco_before_3p", "stock_before_3", "stock_after_3",
                "reco_before_po", "stock_before_po", "stock_after_po",
                "eop_first",
            ]
            for col in diag_cols:
                if col in df_out.columns:
                    df_out[col] = df_out[col].where(pd.notna(df_out[col]), "–")

            df_out.drop(columns=["plan_sales_per_day", "stock_ff", "stock_mp"], errors="ignore", inplace=True)
            df_out = _order_columns(df_out)
            df_out = df_out.reindex(columns=[c for c in _ORDER if c in df_out.columns])
        df_out = df_out.rename(columns=RECOMMENDATION_COLUMN_ALIASES)
        # Пишем «Рекомендации»: добавляем служебную колонку порога, она уйдёт в скрытую
        df_out.to_excel(w, sheet_name="Рекомендации", index=False)
        ws_recs = w.book["Рекомендации"]
        _apply_formats_localized(ws_recs)
        ws_recs.row_dimensions[1].height = 32
        ws_recs.freeze_panes = "A2"

        # 3) Пишем скрытый лист Log с техполями (без debug_*)
        log_cols = [
            "sku",
            "H_days",
            "demand_H",
            "inbound",
            "coverage",
            "target",
            "shortage",
            "moq_step",
            "order_qty",
            "stock_status",
            "algo_version",
        ]
        log_df = df_rec.reindex(columns=log_cols)
        if log_df.shape[1]:
            log_df.to_excel(w, sheet_name="Log", index=False)
            ws_log = w.book["Log"]
            ws_log.sheet_state = "hidden"

        # 4) Затем переносим прочие исходные листы, раскрывая книгу один раз
        try:
            with pd.ExcelFile(in_buf) as xl:
                skip = {"Recommendations", "Рекомендации", "Log", "Заказ на фабрику"}
                for name in xl.sheet_names:
                    if name in skip:
                        continue
                    new_name = "Ввод данных" if name == "Ввод" else name
                    pd.read_excel(xl, name).to_excel(w, sheet_name=new_name, index=False)
        except Exception:
            pass

    return out_buf.getvalue()

def process_excel(xlsx_bytes: bytes) -> bytes:
    items, trans = read_input(xlsx_bytes)
    recs = calculate(items, trans)
    return build_output(xlsx_bytes, recs)
