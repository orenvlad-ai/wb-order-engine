from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Iterable, Optional, Sequence

from openpyxl import Workbook
from openpyxl.styles import Font

from .config import ALGO_VERSION


RECOMMENDATION_HEADERS = [
    "Артикул",
    "Горизонт планирования",
    "Спрос",
    "В пути",
    "Покрытие, шт",
    "Цель, шт",
    "Дефицит, шт",
    "Кратность",
    "Реком. заказ, шт",
    "Снизить план до",
    "Комментарий",
    "Алго",
]

LOG_HEADERS = [
    "generated_at",
    "algo_version",
    "sku_count",
    "in_transit_count",
    "total_volume_units",
    "stocks_info",
]


def _auto_width(worksheet, minimum: int = 10, maximum: int = 50) -> None:
    for column_cells in worksheet.columns:
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            max_len = max(max_len, len(str(cell.value)))
        width = max(minimum, max_len + 2)
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(width, maximum)


def recommendations_to_excel(
    recs: Iterable,
    *,
    sku_count: Optional[int] = None,
    in_transit_count: int = 0,
    total_volume: Optional[int] = None,
    log_items: Optional[Sequence] = None,
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Planner_Recommendations"

    header_font = Font(bold=True)
    ws.append(RECOMMENDATION_HEADERS)
    for cell in ws[1]:
        cell.font = header_font

    recs_list = list(recs)
    for r in recs_list:
        ws.append([
            getattr(r, "sku", None),
            getattr(r, "H_days", None),
            getattr(r, "demand_H", None),
            getattr(r, "inbound", None),
            getattr(r, "coverage", None),
            getattr(r, "target", None),
            getattr(r, "shortage", None),
            getattr(r, "moq_step", None),
            getattr(r, "order_qty", None),
            getattr(r, "reduce_plan_to", None),
            getattr(r, "comment", None),
            getattr(r, "algo_version", ALGO_VERSION),
        ])

    _auto_width(ws, minimum=10, maximum=40)

    log_ws = wb.create_sheet("Log")
    log_ws.append(LOG_HEADERS)
    for cell in log_ws[1]:
        cell.font = header_font

    generated_at = datetime.now().isoformat()
    sku_total = sku_count if sku_count is not None else len(recs_list)
    if total_volume is None:
        total_volume = sum((getattr(r, "order_qty", 0) or 0) for r in recs_list)

    stocks_info = ""
    if log_items:
        parts = []
        for item in log_items:
            sku = getattr(item, "sku", "")
            ff = getattr(item, "safety_stock_ff", None)
            mp = getattr(item, "safety_stock_mp", None)
            if ff is None and mp is None:
                continue
            stocks = f"ff={ff}, mp={mp}"
            parts.append(f"{sku}: {stocks}" if sku else stocks)
        stocks_info = "; ".join(parts)

    log_ws.append([
        generated_at,
        ALGO_VERSION,
        sku_total,
        in_transit_count,
        total_volume,
        stocks_info,
    ])
    _auto_width(log_ws, minimum=15, maximum=50)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
