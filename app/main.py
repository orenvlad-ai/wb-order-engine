# app/main.py
from datetime import date, datetime
from io import BytesIO
from typing import List, Tuple, Dict, Optional

from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse, JSONResponse
from fastapi.templating import Jinja2Templates
from fastapi.staticfiles import StaticFiles

from openpyxl import load_workbook

from engine.calc import calculate
from engine.models import SkuInput, InTransitItem
from engine.config import ALGO_VERSION
from engine.excel import recommendations_to_excel
from adapters.excel_io import (
    BadTemplateError,
    build_output,
    generate_input_template,
    read_input,
)
import uvicorn
import logging


app = FastAPI(title="WB Order Engine")

# --- Шаблоны и статика ---
templates = Jinja2Templates(directory="app/templates")
app.mount("/static", StaticFiles(directory="app/static"), name="static")

# --- UI: форма ввода ---
@app.get("/", response_class=HTMLResponse)
async def input_form(request: Request):
    return templates.TemplateResponse(
        "input_form.html",
        {"request": request, "algo_version": ALGO_VERSION},
    )


@app.get("/health")
async def health_check():
    return {"status": "ok", "algo_version": ALGO_VERSION}


# ---------------------- Скачивание шаблона ----------------------
@app.get("/download_template")
async def download_template():
    buffer = generate_input_template()
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Input_Template.xlsx"'},
    )


@app.get("/download_input_template")
async def download_input_template():
    buffer = generate_input_template()
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="Input_Template.xlsx"'},
    )


# ---------------------- Вспомогательные функции ----------------------
INPUT_REQUIRED_COLS = ["sku", "stock_ff", "stock_mp", "plan_sales_per_day"]
INPUT_OPTIONAL_COLS = ["safety_stock_ff", "safety_stock_mp"]
SETTINGS_REQUIRED_COLS = [
    "prod_lead_time_days",
    "lead_time_cn_msk",
    "lead_time_msk_mp",
    "moq_step_default",
    "safety_stock_ff_default",
    "safety_stock_mp_default",
]
SETTINGS_SHEET_NAME = "Настройки заказа"
INTRANSIT_COLS = ["sku", "qty", "eta_cn_msk"]


def _coerce_date(v) -> date:
    if isinstance(v, date):
        return v
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, str):
        return datetime.strptime(v.strip(), "%Y-%m-%d").date()
    raise ValueError(f"Некорректная дата: {v!r}")


def _is_blank(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def _parse_required_int(value, *, sheet: str, column: str, sku: Optional[str] = None) -> int:
    target = f" для SKU '{sku}'" if sku else ""
    if _is_blank(value):
        raise HTTPException(
            status_code=400,
            detail=f"Лист '{sheet}': колонка '{column}'{target} не заполнена.",
        )
    try:
        return int(value)
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail=f"Лист '{sheet}': колонка '{column}'{target} должна содержать целое число.",
        )


def _parse_required_float(value, *, sheet: str, column: str, sku: str) -> float:
    if _is_blank(value):
        raise HTTPException(
            status_code=400,
            detail=f"Лист '{sheet}': колонка '{column}' для SKU '{sku}' не заполнена.",
        )
    try:
        return float(value)
    except (TypeError, ValueError):
        raise HTTPException(
            status_code=400,
            detail=f"Лист '{sheet}': колонка '{column}' для SKU '{sku}' должна содержать число.",
        )


def _read_settings(ws) -> Dict[str, int]:
    # Считываем первую заполненную строку с общими настройками заказа
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in SETTINGS_REQUIRED_COLS if c not in headers]
    if missing:
        missing_str = ", ".join(missing)
        raise HTTPException(
            status_code=400,
            detail=f"На листе '{SETTINGS_SHEET_NAME}' отсутствуют колонки: {missing_str}.",
        )

    idx = {name: headers.index(name) for name in SETTINGS_REQUIRED_COLS}
    row_values = None
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and not all(_is_blank(v) for v in row):
            row_values = row
            break
    if row_values is None:
        raise HTTPException(
            status_code=400,
            detail="Лист 'Настройки заказа' пуст. Заполни строку с параметрами по умолчанию.",
        )

    settings: Dict[str, int] = {}
    for name in SETTINGS_REQUIRED_COLS:
        value = row_values[idx[name]] if idx[name] < len(row_values) else None
        settings[name] = _parse_required_int(
            value, sheet=SETTINGS_SHEET_NAME, column=name
        )
    return settings


def _read_items(ws, defaults: Dict[str, int]) -> List[SkuInput]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in INPUT_REQUIRED_COLS if c not in headers]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Лист 'Input': нет колонок: {', '.join(missing)}",
        )

    idx = {name: headers.index(name) for name in INPUT_REQUIRED_COLS}
    opt_idx = {name: headers.index(name) for name in INPUT_OPTIONAL_COLS if name in headers}
    items: List[SkuInput] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(_is_blank(v) for v in row):
            continue

        sku_raw = row[idx["sku"]] if idx["sku"] < len(row) else None
        sku = str(sku_raw).strip() if sku_raw is not None else ""
        if not sku:
            raise HTTPException(status_code=400, detail="Лист 'Input': найдена строка без SKU.")

        def get_required(name: str, parser):
            value = row[idx[name]] if idx[name] < len(row) else None
            return parser(value, sheet="Input", column=name, sku=sku)

        stock_ff = get_required("stock_ff", _parse_required_int)
        stock_mp = get_required("stock_mp", _parse_required_int)
        plan_sales_per_day = _parse_required_float(
            row[idx["plan_sales_per_day"]] if idx["plan_sales_per_day"] < len(row) else None,
            sheet="Input",
            column="plan_sales_per_day",
            sku=sku,
        )

        def get_override(name: str, default_value: int) -> int:
            position = opt_idx.get(name)
            if position is None or position >= len(row):
                return default_value
            value = row[position]
            if _is_blank(value):
                return default_value  # пусто → используем дефолт из настроек
            return _parse_required_int(value, sheet="Input", column=name, sku=sku)

        safety_stock_ff = get_override("safety_stock_ff", defaults["safety_stock_ff_default"])
        safety_stock_mp = get_override("safety_stock_mp", defaults["safety_stock_mp_default"])

        items.append(SkuInput(
            sku=sku,
            stock_ff=stock_ff,
            stock_mp=stock_mp,
            plan_sales_per_day=plan_sales_per_day,
            prod_lead_time_days=defaults["prod_lead_time_days"],
            lead_time_cn_msk=defaults["lead_time_cn_msk"],
            lead_time_msk_mp=defaults["lead_time_msk_mp"],
            safety_stock_ff=safety_stock_ff,
            safety_stock_mp=safety_stock_mp,
            moq_step=defaults["moq_step_default"],
        ))
    if not items:
        raise HTTPException(status_code=400, detail="Лист 'Input' пуст.")
    return items


def _read_intransit(ws) -> List[InTransitItem]:
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    missing = [c for c in INTRANSIT_COLS if c not in headers]
    if missing:
        raise HTTPException(status_code=400, detail=f"Лист 'InTransit': нет колонок: {', '.join(missing)}")

    idx = {name: headers.index(name) for name in INTRANSIT_COLS}
    rows: List[InTransitItem] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(_is_blank(v) for v in row):
            continue

        def get(name):
            j = idx[name]
            return row[j]

        try:
            rows.append(InTransitItem(
                sku=str(get("sku")).strip(),
                qty=int(get("qty") or 0),
                eta_cn_msk=_coerce_date(get("eta_cn_msk")),
            ))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Лист 'InTransit': ошибка парсинга строки {row}: {e}")
    return rows


def _parse_input_excel(content: bytes) -> Tuple[List[SkuInput], List[InTransitItem]]:
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Не удалось открыть Excel. Убедись, что это .xlsx файл.")

    if "Input" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="В файле отсутствует лист 'Input'.")
    if SETTINGS_SHEET_NAME not in wb.sheetnames:
        raise HTTPException(
            status_code=400,
            detail="В файле отсутствует лист 'Настройки заказа'.",
        )

    defaults = _read_settings(wb[SETTINGS_SHEET_NAME])  # дефолты применяем ко всем SKU
    items = _read_items(wb["Input"], defaults)
    in_transit: List[InTransitItem] = []
    if "InTransit" in wb.sheetnames:
        in_transit = _read_intransit(wb["InTransit"])
    return items, in_transit


def _excel_from_recs(recs, *, sku_count: int, in_transit_count: int, items: List[SkuInput]):
    recs_list = list(recs)
    total_volume = sum((getattr(r, "order_qty", 0) or 0) for r in recs_list)

    return recommendations_to_excel(
        recs_list,
        sku_count=sku_count,
        in_transit_count=in_transit_count,
        total_volume=total_volume,
        log_items=items,
    )


# ---------------------- Ручной ввод -> Excel ----------------------
@app.post("/calc_excel")
async def calc_excel(
    sku: str = Form(...),
    stock_ff: int = Form(...),
    stock_mp: int = Form(...),
    plan_sales_per_day: float = Form(...),
    prod_lead_time_days: int = Form(...),
    lead_time_cn_msk: int = Form(...),
    lead_time_msk_mp: int = Form(...),
    safety_stock_ff: int = Form(...),
    safety_stock_mp: int = Form(...),
    moq_step: int = Form(...),

    in_transit_qty: List[int] = Form(default=[]),
    in_transit_eta: List[str] = Form(default=[]),
):
    item = SkuInput(
        sku=sku.strip(),
        stock_ff=stock_ff,
        stock_mp=stock_mp,
        plan_sales_per_day=plan_sales_per_day,
        prod_lead_time_days=prod_lead_time_days,
        lead_time_cn_msk=lead_time_cn_msk,
        lead_time_msk_mp=lead_time_msk_mp,
        safety_stock_ff=safety_stock_ff,
        safety_stock_mp=safety_stock_mp,
        moq_step=moq_step,
    )

    in_transit: List[InTransitItem] = []
    for q, eta_str in zip(in_transit_qty or [], in_transit_eta or []):
        if not eta_str:
            continue
        eta = datetime.strptime(eta_str, "%Y-%m-%d").date()
        in_transit.append(InTransitItem(sku=item.sku, qty=int(q), eta_cn_msk=eta))

    recs = calculate([item], in_transit=in_transit)
    buff = _excel_from_recs(
        recs,
        sku_count=len(recs),
        in_transit_count=len(in_transit),
        items=[item],
    )

    filename = f"Planner_Recommendations_{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buff,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


# ---------------------- Загрузка Excel -> Excel ----------------------
@app.post("/upload_excel")
async def upload_excel(file: UploadFile = File(...)):
    try:
        if not file.filename.lower().endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Ожидается .xlsx файл.")

        content = await file.read()
        items, in_transit = read_input(content)
        recs = calculate(items, in_transit)
        out_bytes = build_output(content, recs)

        fname = f"Planner_Recommendations_{date.today().isoformat()}.xlsx"
        return StreamingResponse(
            BytesIO(out_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'}
        )
    except HTTPException as exc:
        return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})
    except BadTemplateError as exc:
        return JSONResponse(status_code=400, content={"detail": str(exc)})
    except Exception:
        logging.exception("Unexpected error while processing Excel upload")
        return JSONResponse(
            status_code=500,
            content={"detail": "Внутренняя ошибка при обработке Excel"}
        )


# ---------------------- Локальный запуск ----------------------
if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=8000, reload=True)
